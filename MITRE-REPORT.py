import json
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side, BORDER_THICK

#prompt = input('Enter path to Mitre Dataset json file:')
prompt = ".\\Mitre_Dataset.json"
bAtomic = ""
while bAtomic != "y" and bAtomic != "n":
    bAtomic = input("Would you like to ingest an atomic test results dataset?(y,n)")
if bAtomic == "y":
    bAtomic = True
else:
    bAtomic = False


if bAtomic:
    prompt2 = input('Enter path to Mitre Dataset with test data added:')
bThreatGroup = ""
fThreatGroup = ".\\ThreatGroups.json"
while bThreatGroup != "y" and bThreatGroup != "n":
    bThreatGroup = input("Would you like to overlay Data Sources with a known threat group?(y,n)")
if bThreatGroup == "y":
    bThreatGroup = True
else:
    bThreatGroup = False
#data_components = list(input("Enter a list of data components i.e ['Command Execution', 'Host Status', 'Windows Registry Key Modification']"))
data_components = ["Script Execution", "File Access", "File Creation", "File Deletion", "File Metadata", "File Modification", "Logon Session Creation", "Logon Session Metadata", "Malware Content", "Malware Metadata", "Network Connection Creation", "Network Traffic Content", "Network Traffic Flow", "OS API Execution", "Process Access", "Process Creation", "Process Metadata", "Process Modification", "Process Termination", "Network Share Access", "Windows Registry Key Access", "Windows Registry Key Creation", "Windows Registry Key Deletion", "Windows Registry Key Modification"]
#output_file = input("Enter the name of the output file")

# load the indicated files from json into a python dictionary
f = open(prompt)
data = json.load(f)
f.close()

if bAtomic:
    f2 = open(prompt2)
    test_data = json.load(f2)
    f2.close()

if bThreatGroup:
    f3 = open(".\\ThreatGroups.json")
    threat_group_data = json.load(f3)
    f3.close()
    group = ""
    while group == "":
        group = input("Enter the name of the threat group?")
        if group in threat_group_data:
            continue
        else:
            group = ""

#create workbook and add sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "DataSrc Coverage"
if bAtomic:
    test_sheet = workbook.create_sheet()
    test_sheet.title = "Atomic Test Coverage"
if bThreatGroup:
    overlay_sheet = workbook.create_sheet()
    overlay_sheet.title = group + "Overlay"

# EXCEL OUTPUT SETTINGS
red = Color(indexed=2)
green = Color(indexed=3)
yellow = Color(indexed=5)
blue = Color(indexed=35)


thick_border = Border(
    left=Side(border_style=BORDER_THICK, color='00000000'),
    right=Side(border_style=BORDER_THICK, color='00000000'),
    top=Side(border_style=BORDER_THICK, color='00000000'),
    bottom=Side(border_style=BORDER_THICK, color='00000000')
)

thick_left = Border(
    left=Side(border_style=BORDER_THICK, color='00000000')
)

thick_right = Border(
    right=Side(border_style=BORDER_THICK, color='00000000')
)



tactics = [i for i in data]

def center_cell(row, col):
    cell = sheet.cell(row=row, column=col)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# draw tactics on first row
def excel_tactic_setup(excel_sheet):
    column1 = None
    column2 = None
    # merge top columns into groups of 4
    for i in range(len(data) * 6 + 1):
        if (i != 0): # TURN ON THICK BORDERS FOR TACTICS HERE
            cell = excel_sheet[get_column_letter(i) + str(1)]
            cell.border = thick_border
        if (i % 6 == 0 or i == 0) and column1 == None:
            if i == 0:
                i = 1
                #column2 = get_column_letter(6)
            column1 = get_column_letter(i)
            col = i
        elif i % 6 == 0:
            column2 = get_column_letter(i)
        if column1 and column2:
            excel_sheet.merge_cells(column1 + str(1) + ":" + column2 + str(1))
            # write title
            cell = excel_sheet[column1 + str(1)]
            if i <= 6:
                cell.value = tactics[0]
            else:
                try:
                    cell.value = tactics[int(col / 6)]
                except Exception:
                    continue
            cell.alignment = Alignment(horizontal='center', vertical='center')
            column1 = get_column_letter(i+1)
            col = i + 1
            column2 = None


########## DATA SOURCE COVERAGE SHEET FUNCTIONS  ##############

# technique, subtechnique, data source across second row
def sub_technique_setup():
    column1 = None
    column2 = None
    bWrote1 = False
    bWrote2 = False
    bWrote3 = False
    for i in range(len(data) * 6 + 1):
        if (i != 0): # TURN ON THICK BORDERS FOR CATEGORIES HERE
            cell = sheet[get_column_letter(i) + str(2)]
            cell.border = thick_border
        if not column1:
            column1 = get_column_letter(1)
        elif i % 2 == 0 and i > 1:
            column2 = get_column_letter(i)
        if column1 and column2 and i < 92:
            sheet.merge_cells(column1 + str(2) + ":" + column2 + str(2))
            cell = sheet[column1 + str(2)]
            if i % 2 == 0 and i < 96 and bWrote1 and bWrote2:
                cell.value = "Data Source"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote3 = True
                if bWrote1 and bWrote2 and bWrote3:
                    bWrote1 = False
                    bWrote2 = False
                    bWrote3 = False
            elif i % 2 == 0 and i < 92 and bWrote1:
                cell.value = "Sub-Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote2 = True
            elif i % 2 == 0 and i < 90 and not bWrote2 and not bWrote3:
                cell.value = "Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote1 = True



def draw_datasrc_coverage(tactic, sect):
    # find base cols and rows
    row = 3
    for i in range(len(sect) * 6):
        if i > 0:
            cell = sheet[get_column_letter(i) + "1"]
            if cell.value == tactic:
                col = i
                continue
    tactic_dict = sect[tactic]

    for tech in tactic_dict:
        sheet.merge_cells(get_column_letter(col) + str(row) + ":" + get_column_letter(col + 1) + str(row))
        cell = sheet[get_column_letter(col) + str(row)]
        cell.value = tech
        cell.border = thick_left
        new_cell = sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right
        row += 1
        cell = sheet[get_column_letter(col) + str(row)]
        cell.border = thick_left
        new_cell = sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right
        for sub_tech in tactic_dict[tech]:
            if sub_tech == "Data Sources":
                for src in tactic_dict[tech][sub_tech]:
                    for s in tactic_dict[tech][sub_tech][src]:
                        used_column = get_column_letter(col + 4)
                        used_column2 = get_column_letter(int(col + 5))
                        cell = sheet[used_column + str(row)]
                        cell.value = s
                        sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                        row += 1
                        if tactic_dict[tech][sub_tech][src][s] != "COVERED":
                            cell.fill = PatternFill("solid", start_color=red)
                        cell = sheet[get_column_letter(col) + str(row)]
                        cell.border = thick_left
                        new_cell = sheet[get_column_letter(col + 5) + str(row)]
                        new_cell.border = thick_right
            else:
                    used_column = get_column_letter(int(col + 2))
                    used_column2 = get_column_letter(int(col + 3))
                    sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                    cell = sheet[used_column + str(row)]
                    cell.value = sub_tech
                    row += 1
                    new_cell = sheet[get_column_letter(col) + str(row)]
                    new_cell.border = thick_left
                    new_cell = sheet[get_column_letter(col + 5) + str(row)]
                    new_cell.border = thick_right
                    for data_src in tactic_dict[tech][sub_tech]:
                        for data in tactic_dict[tech][sub_tech][data_src]:
                            for comp in tactic_dict[tech][sub_tech][data_src][data]:
                                used_column = get_column_letter(int(col + 4))
                                used_column2 = get_column_letter(int(col + 5))
                                cell = sheet[used_column + str(row)]
                                cell.value = comp
                                sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                                row += 1
                                # change color of data sources that are not being analyzed
                                if tactic_dict[tech][sub_tech][data_src][data][comp] != "COVERED":
                                    cell.fill = PatternFill("solid", start_color=red)
                                cell = sheet[get_column_letter(col) + str(row)]
                                cell.border = thick_left
                                new_cell = sheet[get_column_letter(col + 5) + str(row)]
                                new_cell.border = thick_right
########## DATA SOURCE COVERAGE SHEET FUNCTIONS  ##############



########## ATOMIC TEST RESULTS SHEET FUNCTIONS #############
def color_cell(cell):
    string = cell.value
    if "/" not in string:
        return
    temp = string.split("/")
    num1 = int(temp[0])
    num2 = int(temp[1])
    pct = num1 / num2
    if pct < 0.25:
        cell.fill = PatternFill("solid", start_color=red)
    elif .25 < pct < .75:
        cell.fill = PatternFill("solid", start_color=yellow)
    elif pct > .75:
        cell.fill = PatternFill("solid", start_color=green)



def sub_technique_tests_setup(): # run sub category set up for atomic tests sheet
    column1 = None
    column2 = None
    bWrote1 = False
    bWrote2 = False
    bWrote3 = False
    for i in range(len(data) * 6 + 1):
        if (i != 0): # TURN ON THICK BORDERS FOR CATEGORIES HERE
            cell = test_sheet[get_column_letter(i) + str(2)]
            cell.border = thick_border
        if not column1:
            column1 = get_column_letter(1)
        elif i % 2 == 0 and i > 1:
            column2 = get_column_letter(i)
        if column1 and column2 and i < 92:
            test_sheet.merge_cells(column1 + str(2) + ":" + column2 + str(2))
            cell = test_sheet[column1 + str(2)]
            if i % 2 == 0 and i < 96 and bWrote1 and bWrote2:
                cell.value = "Test Results"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote3 = True
                if bWrote1 and bWrote2 and bWrote3:
                    bWrote1 = False
                    bWrote2 = False
                    bWrote3 = False
            elif i % 2 == 0 and i < 92 and bWrote1:
                cell.value = "Sub-Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote2 = True
            elif i % 2 == 0 and i < 90 and not bWrote2 and not bWrote3:
                cell.value = "Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote1 = True

def draw_test_coverage(tactic, sect):
    # find base cols and rows
    row = 3
    for i in range(len(sect) * 6):
        if i > 0:
            cell = test_sheet[get_column_letter(i) + "1"]
            if cell.value == tactic:
                col = i
                continue
    tactic_dict = sect[tactic]

    for tech in tactic_dict:
        test_sheet.merge_cells(get_column_letter(col) + str(row) + ":" + get_column_letter(col + 1) + str(row))
        cell = test_sheet[get_column_letter(col) + str(row)]
        cell.value = tech
        cell = test_sheet[get_column_letter(col + 4) + str(row + 1)]
        test_sheet.merge_cells(get_column_letter(col+4) + str(row + 1) + ":" + get_column_letter(col+5) + str(row + 1))
        cell.value = "0/0"
        cell.alignment = Alignment(horizontal='center', vertical='center')

        new_cell = test_sheet[get_column_letter(col) + str(row)]
        new_cell.border = thick_left
        new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right

        row += 1

        new_cell = test_sheet[get_column_letter(col) + str(row)]
        new_cell.border = thick_left
        new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right
        try:
            if tactic_dict[tech]["Test Results"]:
                used_column = get_column_letter(col + 4)
                used_column2 = get_column_letter(int(col + 5))
                cell = test_sheet[used_column + str(row)]
                try:
                    cell.value = tactic_dict[tech]["Test Results"]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    test_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                    row += 1
                    new_cell = test_sheet[get_column_letter(col) + str(row)]
                    new_cell.border = thick_left
                    new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
                    new_cell.border = thick_right
                    color_cell(cell)
                except Exception:
                    cell.value = "0/0"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    test_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                    row += 1
                    new_cell = test_sheet[get_column_letter(col) + str(row)]
                    new_cell.border = thick_left
                    new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
                    new_cell.border = thick_right
        except Exception as e:
            row += 1
            new_cell = test_sheet[get_column_letter(col) + str(row)]
            new_cell.border = thick_left
            new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
            new_cell.border = thick_right
            pass
        for sub_tech in tactic_dict[tech]:
            if sub_tech == "Data Sources" or sub_tech == "Test Results":
                pass

            else:
                    used_column = get_column_letter(int(col + 2))
                    used_column2 = get_column_letter(int(col + 3))
                    test_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                    cell = test_sheet[used_column + str(row)]
                    cell.value = sub_tech
                    row += 1
                    new_cell = test_sheet[get_column_letter(col) + str(row)]
                    new_cell.border = thick_left
                    new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
                    new_cell.border = thick_right
                    for data_src in tactic_dict[tech][sub_tech]:
                        if data_src != "Test Results":
                            used_column = get_column_letter(int(col + 4))
                            used_column2 = get_column_letter(int(col + 5))
                            cell = test_sheet[used_column + str(row)]
                            try:
                                cell.value = tactic_dict[tech][sub_tech]["Test Results"]
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                test_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                                # change color of data sources that are not being analyzed
                                color_cell(cell)
                                row += 1
                                new_cell = test_sheet[get_column_letter(col) + str(row)]
                                new_cell.border = thick_left
                                new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
                                new_cell.border = thick_right
                            except:
                                cell.value = "0/0"
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                test_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                                row += 1
                                new_cell = test_sheet[get_column_letter(col) + str(row)]
                                new_cell.border = thick_left
                                new_cell = test_sheet[get_column_letter(col + 5) + str(row)]
                                new_cell.border = thick_right


########## ATOMIC TEST RESULTS SHEET FUNCTIONS #############



########## THREAT GROUP OVERLAY SHEET FUNCTIONS #############

def sub_technique_overlay_setup(): # run sub category set up for atomic tests sheet
    column1 = None
    column2 = None
    bWrote1 = False
    bWrote2 = False
    bWrote3 = False
    for i in range(len(data) * 6 + 1):
        if (i != 0): # TURN ON THICK BORDERS FOR CATEGORIES HERE
            cell = overlay_sheet[get_column_letter(i) + str(2)]
            cell.border = thick_border
        if not column1:
            column1 = get_column_letter(1)
        elif i % 2 == 0 and i > 1:
            column2 = get_column_letter(i)
        if column1 and column2 and i < 92:
            overlay_sheet.merge_cells(column1 + str(2) + ":" + column2 + str(2))
            cell = overlay_sheet[column1 + str(2)]
            if i % 2 == 0 and i < 96 and bWrote1 and bWrote2:
                cell.value = "Data Source"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote3 = True
                if bWrote1 and bWrote2 and bWrote3:
                    bWrote1 = False
                    bWrote2 = False
                    bWrote3 = False
            elif i % 2 == 0 and i < 92 and bWrote1:
                cell.value = "Sub-Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote2 = True
            elif i % 2 == 0 and i < 90 and not bWrote2 and not bWrote3:
                cell.value = "Technique"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                column1 = get_column_letter(i + 1)
                column2 = None
                bWrote1 = True


def draw_overlay_coverage(tactic, sect):
    # find base cols and rows
    row = 3
    for i in range(len(sect) * 6):
        if i > 0:
            cell = overlay_sheet[get_column_letter(i) + "1"]
            if cell.value == tactic:
                col = i
                continue
    tactic_dict = sect[tactic]

    for tech in tactic_dict:
        overlay_sheet.merge_cells(get_column_letter(col) + str(row) + ":" + get_column_letter(col + 1) + str(row))
        tech_cell = overlay_sheet[get_column_letter(col) + str(row)]
        tech_cell.value = tech
        new_cell = overlay_sheet[get_column_letter(col) + str(row)]
        new_cell.border = thick_left
        new_cell = overlay_sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right
        row += 1
        new_cell = overlay_sheet[get_column_letter(col) + str(row)]
        new_cell.border = thick_left
        new_cell = overlay_sheet[get_column_letter(col + 5) + str(row)]
        new_cell.border = thick_right
        for sub_tech in tactic_dict[tech]:
            if sub_tech == "Data Sources":
                for src in tactic_dict[tech][sub_tech]:
                    for s in tactic_dict[tech][sub_tech][src]:
                        used_column = get_column_letter(col + 4)
                        used_column2 = get_column_letter(int(col + 5))
                        cell = overlay_sheet[used_column + str(row)]
                        cell.value = s
                        overlay_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                        row += 1
                        new_cell = overlay_sheet[get_column_letter(col) + str(row)]
                        new_cell.border = thick_left
                        new_cell = overlay_sheet[get_column_letter(col + 5) + str(row)]
                        new_cell.border = thick_right
                        try:
                            if tactic_dict[tech][sub_tech][src][s] != "COVERED" and s in threat_group_data[group][tech]["Data Sources"]:
                                cell.fill = PatternFill("solid", start_color=red)
                                tech_cell.fill = PatternFill("solid", start_color=blue)
                            elif tactic_dict[tech][sub_tech][src][s] == "COVERED" and s in threat_group_data[group][tech]["Data Sources"]:
                                cell.fill = PatternFill("solid", start_color=green)
                                tech_cell.fill = PatternFill("solid", start_color=blue)
                        except Exception:
                            pass
            else:
                    used_column = get_column_letter(int(col + 2))
                    used_column2 = get_column_letter(int(col + 3))
                    overlay_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                    sub_cell = overlay_sheet[used_column + str(row)]
                    sub_cell.value = sub_tech
                    row += 1
                    new_cell = overlay_sheet[get_column_letter(col) + str(row)]
                    new_cell.border = thick_left
                    new_cell = overlay_sheet[get_column_letter(col + 5) + str(row)]
                    new_cell.border = thick_right
                    for data_src in tactic_dict[tech][sub_tech]:
                        for data in tactic_dict[tech][sub_tech][data_src]:
                            for comp in tactic_dict[tech][sub_tech][data_src][data]:
                                used_column = get_column_letter(int(col + 4))
                                used_column2 = get_column_letter(int(col + 5))
                                cell = overlay_sheet[used_column + str(row)]
                                cell.value = comp
                                overlay_sheet.merge_cells(used_column + str(row) + ":" + used_column2 + str(row))
                                row += 1
                                new_cell = overlay_sheet[get_column_letter(col) + str(row)]
                                new_cell.border = thick_left
                                new_cell = overlay_sheet[get_column_letter(col + 5) + str(row)]
                                new_cell.border = thick_right
                                # change color of data sources that are not being analyzed
                                try:
                                    if tactic_dict[tech][sub_tech][data_src][data][comp] != "COVERED" and comp in threat_group_data[group][tech][sub_tech]["Data Sources"]:
                                        cell.fill = PatternFill("solid", start_color=red)
                                        sub_cell.fill = PatternFill("solid", start_color=blue)
                                    elif tactic_dict[tech][sub_tech][data_src][data][comp] == "COVERED" and comp in threat_group_data[group][tech][sub_tech]["Data Sources"]:
                                        cell.fill = PatternFill("solid", start_color=green)
                                        sub_cell.fill = PatternFill("solid", start_color=blue)
                                except Exception as e:
                                    pass
########## THREAT GROUP OVERLAY SHEET FUNCTIONS #############








####################Nested FOR loop to traverse loaded JSON file and mark the correct data sources as covered##########################
for tactic in data:
    for tactic in data:
        for technique in data[tactic]:
            for sub_technique in data[tactic][technique]:
                if sub_technique == "Data Sources":
                    pass
                    for source in data[tactic][technique][sub_technique]:
                        for component in data[tactic][technique][sub_technique][source]:
                            if component in data_components:
                                data[tactic][technique][sub_technique][source][component] = "COVERED"
                else:
                    for data_sources in data[tactic][technique][sub_technique]:
                        for source in data[tactic][technique][sub_technique][data_sources]:
                            for component in data[tactic][technique][sub_technique][data_sources][source]:
                                if component in data_components:
                                    data[tactic][technique][sub_technique][data_sources][source][component] = "COVERED"
#######################################################################################################################################






print(r"""
____   ____.__                    .__  ____                     
\   \ /   /|__|_________ _______  |  |/_   |_______ ___________ 
 \   Y   / |  \___   /  |  \__  \ |  | |   \___   // __ \_  __ \
  \     /  |  |/    /|  |  // __ \|  |_|   |/    /\  ___/|  | \/
   \___/   |__/_____ \____/(____  /____/___/_____ \\___  >__|   
                    \/          \/               \/    \/       
                    dev Cole Strickler
""")



########DRAWING DATASOURCE COVERAGE###########
excel_tactic_setup(excel_sheet=sheet)
sub_technique_setup()
for tactic in tactics:
    draw_datasrc_coverage(tactic, data)
########DRAWING DATASOURCE COVERAGE###########



########DRAWING ATOMIC TEST RESULTS###########
if bAtomic:
    excel_tactic_setup(excel_sheet=test_sheet)
    sub_technique_tests_setup()
    for tactic in tactics:
        draw_test_coverage(tactic, test_data)
########DRAWING ATOMIC TEST RESULTS###########


###########DRAWING THREAT OVERLAY#############
if bThreatGroup:
    excel_tactic_setup(excel_sheet=overlay_sheet)
    sub_technique_overlay_setup() # need to make a new function for this sheet
    for tactic in tactics:
        draw_overlay_coverage(tactic, data)

###########DRAWING THREAT OVERLAY#############







workbook.save("MITRE-REPORT.xlsx")

print("\n\nFINISHED. Saved to MITRE-REPORT.xlsx")